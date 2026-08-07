import os
import shutil
import json
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import load_workbook

# Paths
ROOT = Path("E:/OneDrive - Uncle Robert Consulting LLC/Working Docs/AI Native Agency Deepened/AgentLab")
EXCEL_PATH = Path("E:/OneDrive - Uncle Robert Consulting LLC/Desktop/Uncle Robert Records.xlsx")
TEMP_EXCEL = ROOT / "temp_records.xlsx"
HTML_PATH = Path("E:/OneDrive - Uncle Robert Consulting LLC/Desktop/command-center-html.html")

def copy_excel_safely():
    try:
        import subprocess
        # Use Node's proven sharing-flag copy to bypass Excel lock
        cmd = f"node -e \"import {{ copyFile }} from 'node:fs/promises'; await copyFile('{EXCEL_PATH.as_posix()}', '{TEMP_EXCEL.as_posix()}');\""
        subprocess.run(cmd, shell=True, check=True, capture_output=True)
        return True
    except Exception as e:
        print(f"Error copying Excel: {e}")
        return False

def parse_app_builds(ws):
    apps = []
    current_app = None
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val is None:
            continue
        val = str(val).strip()
        if not val:
            continue
        
        if val.startswith("http://") or val.startswith("https://"):
            if current_app:
                current_app["urls"].append(val)
        elif val.lower() == "built into website":
            if current_app:
                current_app["notes"].append(val)
        elif "progress reports" in val.lower() or "apps builds" in val.lower():
            continue
        else:
            if current_app:
                apps.append(current_app)
            current_app = {"name": val, "urls": [], "notes": []}
    if current_app:
        apps.append(current_app)
    return apps

def parse_accomplishments(ws):
    sections = []
    current_section = None
    for r in range(1, ws.max_row + 1):
        val_a = ws.cell(row=r, column=1).value
        val_d = ws.cell(row=r, column=4).value
        
        if val_a is None:
            continue
        val_a = str(val_a).strip()
        if not val_a:
            continue
            
        if "corresponds to urc workflow" in val_a.lower():
            if current_section and val_d:
                current_section["workflow"] = str(val_d).strip()
        elif val_a.startswith("http://") or val_a.startswith("https://"):
            if current_section:
                current_section["links"].append(val_a)
        else:
            known_categories = [
                "agent lab", "market marksman app", "pulse app", 
                "bills tracker", "48-hour linkedin authority system", 
                "consulting assessment question generator to app"
            ]
            is_category = any(cat in val_a.lower() for cat in known_categories)
            if is_category:
                if current_section:
                    sections.append(current_section)
                current_section = {"category": val_a, "items": [], "workflow": None, "links": []}
            else:
                if current_section:
                    current_section["items"].append(val_a)
    if current_section:
        sections.append(current_section)
    return sections

def parse_new_accounts(ws):
    accounts = []
    for r in range(3, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        url = ws.cell(row=r, column=2).value
        trial = ws.cell(row=r, column=3).value
        free = ws.cell(row=r, column=4).value
        if name and url:
            accounts.append({
                "name": str(name).strip(),
                "url": str(url).strip(),
                "trial": str(trial).strip() if trial else "N",
                "free": str(free).strip() if free else ""
            })
    return accounts

def parse_lead_list(ws):
    leads = []
    current_source = None
    for r in range(1, ws.max_row + 1):
        val_a = ws.cell(row=r, column=1).value
        val_e = ws.cell(row=r, column=5).value
        
        if val_a is None:
            continue
        val_a = str(val_a).strip()
        if not val_a:
            continue
            
        if "known registrations at" in val_a.lower():
            continue
            
        if ".com" in val_a.lower() or ".online" in val_a.lower() or ".tech" in val_a.lower():
            current_source = val_a
        else:
            lead_date = str(val_e).strip() if val_e else ""
            leads.append({
                "source": current_source or "General",
                "detail": val_a,
                "date": lead_date
            })
    return leads

def parse_social_posts(ws, target_date_num=46245):
    posts = []
    for r in range(2, ws.max_row + 1):
        date_val = ws.cell(row=r, column=1).value
        if date_val is None:
            continue
        try:
            date_num = int(date_val)
        except ValueError:
            continue
        if date_num == target_date_num:
            platform = ws.cell(row=r, column=3).value
            variation = ws.cell(row=r, column=4).value
            copy_text = ws.cell(row=r, column=5).value
            graphic = ws.cell(row=r, column=6).value
            if platform and copy_text:
                posts.append({
                    "platform": str(platform).strip(),
                    "variation": variation,
                    "copy": str(copy_text).strip(),
                    "graphic": str(graphic).strip() if graphic else ""
                })
    return posts

def build_html_sections(apps, accomplishments, accounts, leads, posts):
    html = []
    html.append("\n  <!-- EXCEL RECORDS START -->")
    
    # 1. Upcoming Social Posts Card
    if posts:
        html.append('\n  <h2>📅 Upcoming Social Posts (Tue, Aug 11, 2026 - Day 9, Var 3)</h2>')
        html.append('  <div style="background:#fff; border:1px solid var(--line); border-radius:12px; padding:16px 20px; box-shadow:0 1px 3px rgba(11,36,64,.05); margin-bottom:24px;">')
        html.append('    <p style="font-size:.9rem; color:var(--muted); margin:0 0 16px 0;">📋 <b>Ready to publish!</b> Click "Copy Post Copy" and upload with the suggested graphic file.</p>')
        html.append('    <div style="display:grid; gap:16px; grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));">')
        for post in posts:
            p_color = "#0077b5" if post["platform"].lower() == "linkedin" else "#1877f2" if post["platform"].lower() == "facebook" else "#e1306c" if post["platform"].lower() == "instagram" else "#25d366"
            html.append(f'      <div style="background:var(--bg); border:1px solid var(--line); border-radius:10px; padding:14px; display:flex; flex-direction:column; justify-content:space-between;">')
            html.append(f'        <div>')
            html.append(f'          <div style="font-weight:700; color:{p_color}; font-size:.95rem; margin-bottom:4px; display:flex; align-items:center; justify-content:space-between;">')
            html.append(f'            <span>{post["platform"]} (Var {post["variation"]})</span>')
            if post["graphic"]:
                html.append(f'            <span style="font-size:.75rem; font-weight:600; background:#fff; border:1px solid var(--line); padding:2px 8px; border-radius:999px; color:var(--muted);">{post["graphic"]}</span>')
            html.append(f'          </div>')
            html.append(f'          <textarea style="width:100%; min-height:120px; font-family:inherit; font-size:.85rem; padding:8px; border:1px solid var(--line); border-radius:6px; background:#fff; resize:vertical; color:var(--ink); margin-bottom:8px;" readonly>{post["copy"]}</textarea>')
            html.append(f'        </div>')
            html.append(f'        <button onclick="navigator.clipboard.writeText(this.previousElementSibling.value || this.parentElement.querySelector(\'textarea\').value); showToast(\'📋 {post["platform"]} post copy copied!\');" ')
            html.append(f'                style="width:100%; border:0; background:var(--ink); color:#fff; font-weight:700; padding:8px; border-radius:6px; font-size:.82rem; cursor:pointer; transition:background .1s;">')
            html.append(f'          📋 Copy Post Copy')
            html.append(f'        </button>')
            html.append(f'      </div>')
        html.append('    </div>')
        html.append('  </div>')

    # 2. Live App Builds Section
    if apps:
        html.append('\n  <h2>📱 Live App Builds</h2>')
        html.append('  <div style="display:grid; gap:12px; grid-template-columns: repeat(auto-fit, minmax(240px, 1fr)); margin-bottom:24px;">')
        for app in apps:
            if not app["urls"]:
                continue
            html.append('    <div style="background:#fff; border:1px solid var(--line); border-radius:10px; padding:12px 16px; box-shadow:0 1px 3px rgba(11,36,64,.05); display:flex; flex-direction:column; justify-content:space-between;">')
            html.append('      <div>')
            html.append(f'        <div style="font-weight:700; font-size:.95rem; color:var(--ink); margin-bottom:6px;">🚀 {app["name"]}</div>')
            if app["notes"]:
                notes_str = "; ".join(app["notes"])
                html.append(f'        <div style="font-size:.78rem; color:var(--muted); margin-bottom:10px;">{notes_str}</div>')
            html.append('      </div>')
            html.append('      <div style="display:flex; flex-direction:column; gap:6px; margin-top:8px;">')
            for url in app["urls"]:
                u_label = "🔗 Deploy/Preview" if "run.app" in url or "preview" in url or "base44" in url or "expo" in url else "💻 Replit Repo"
                html.append(f'        <a href="{url}" target="_blank" style="text-decoration:none; text-align:center; font-size:.78rem; font-weight:700; color:var(--teal-dk); background:#eef7f7; border:1px solid #cfeaea; padding:6px 10px; border-radius:6px; transition:all .1s;">{u_label}</a>')
            html.append('      </div>')
            html.append('    </div>')
        html.append('  </div>')

    # 3. Accomplished This Week Section
    if accomplishments:
        html.append('\n  <h2>✅ Accomplished This Week</h2>')
        html.append('  <div style="background:#fff; border:1px solid var(--line); border-radius:12px; padding:16px 20px; box-shadow:0 1px 3px rgba(11,36,64,.05); margin-bottom:24px;">')
        html.append('    <div style="display:flex; flex-direction:column; gap:16px;">')
        for sec in accomplishments:
            html.append('      <div style="border-bottom:1px solid var(--line); padding-bottom:12px; margin-bottom:4px;">')
            html.append('        <div style="display:flex; align-items:center; justify-content:space-between; flex-wrap:wrap; margin-bottom:6px;">')
            html.append(f'          <div style="font-weight:700; font-size:.95rem; color:var(--ink);">🛠️ {sec["category"]}</div>')
            if sec["workflow"]:
                html.append(f'          <div style="font-size:.72rem; font-weight:600; background:#f1f5f9; color:var(--muted); padding:2px 8px; border-radius:4px; border:1px solid var(--line);">{sec["workflow"]}</div>')
            html.append('        </div>')
            html.append('        <ul style="margin:4px 0 0; padding-left:18px; list-style-type:disc;">')
            for item in sec["items"]:
                html.append(f'          <li style="margin-bottom:3px; font-size:.85rem; color:var(--ink);">{item}</li>')
            html.append('        </ul>')
            if sec["links"]:
                html.append('        <div style="display:flex; gap:8px; margin-top:8px; padding-left:18px;">')
                for link in sec["links"]:
                    html.append(f'          <a href="{link}" target="_blank" style="font-size:.75rem; text-decoration:none; color:var(--teal-dk); font-weight:600;">🔗 Link to Resource</a>')
                html.append('        </div>')
            html.append('      </div>')
        html.append('    </div>')
        html.append('  </div>')

    # 4. New Platforms & Accounts + Lead List (Side-by-side)
    html.append('\n  <div style="display:grid; gap:16px; grid-template-columns: repeat(auto-fit, minmax(300px, 1fr)); margin-bottom:24px;">')
    
    # 4a. New Accounts Added This Week Card
    if accounts:
        html.append('    <div style="background:#fff; border:1px solid var(--line); border-radius:12px; padding:16px 20px; box-shadow:0 1px 3px rgba(11,36,64,.05); display:flex; flex-direction:column; justify-content:space-between;">')
        html.append('      <div>')
        html.append('        <h3 style="font-size:1.0rem; margin:0 0 12px 0; font-weight:700; color:var(--ink);">🔑 New Platforms Added This Week</h3>')
        html.append('        <table style="width:100%; border-collapse:collapse; font-size:.82rem;">')
        html.append('          <thead>')
        html.append('            <tr style="border-bottom:1px solid var(--line); text-align:left; color:var(--muted); font-weight:700;">')
        html.append('              <th style="padding:6px 4px;">Platform</th>')
        html.append('              <th style="padding:6px 4px;">Type / Limit</th>')
        html.append('              <th style="padding:6px 4px; text-align:right;">Action</th>')
        html.append('            </tr>')
        html.append('          </thead>')
        html.append('          <tbody>')
        for acct in accounts:
            label = "Trial" if acct["trial"].lower() == "y" else "Free" if "limited" in acct["free"].lower() else "PAYG"
            html.append(f'            <tr style="border-bottom:1px solid rgba(0,0,0,.05);">')
            html.append(f'              <td style="padding:6px 4px; font-weight:600; color:var(--ink);">{acct["name"]}</td>')
            html.append(f'              <td style="padding:6px 4px; color:var(--muted);">{label} ({acct["free"]})</td>')
            html.append(f'              <td style="padding:6px 4px; text-align:right;"><a href="{acct["url"]}" target="_blank" style="text-decoration:none; color:var(--teal-dk); font-weight:700;">Open 🔗</a></td>')
            html.append(f'            </tr>')
        html.append('          </tbody>')
        html.append('        </table>')
        html.append('      </div>')
        html.append('    </div>')

    # 4b. Lead Registrations Card
    if leads:
        html.append('    <div style="background:#fff; border:1px solid var(--line); border-radius:12px; padding:16px 20px; box-shadow:0 1px 3px rgba(11,36,64,.05); display:flex; flex-direction:column; justify-content:space-between;">')
        html.append('      <div>')
        html.append('        <h3 style="font-size:1.0rem; margin:0 0 12px 0; font-weight:700; color:var(--ink);">👥 Recent Lead Registrations</h3>')
        html.append('        <ul style="margin:0; padding-left:0; list-style:none;">')
        for lead in leads:
            html.append('          <li style="border-bottom:1px solid rgba(0,0,0,.05); padding:6px 0; display:flex; justify-content:space-between; align-items:center; font-size:.82rem;">')
            html.append('            <div>')
            html.append(f'              <span style="font-weight:600; color:var(--ink);">{lead["detail"]}</span>')
            html.append(f'              <div style="font-size:.7rem; color:var(--muted);">{lead["source"]}</div>')
            html.append('            </div>')
            if lead["date"]:
                date_str = lead["date"].split("T")[0]
                html.append(f'            <span style="font-size:.7rem; color:var(--muted); background:var(--bg); border:1px solid var(--line); padding:2px 6px; border-radius:4px;">{date_str}</span>')
            html.append('          </li>')
        html.append('        </ul>')
        html.append('      </div>')
        html.append('    </div>')

    html.append('  </div>')
    html.append("\n  <!-- EXCEL RECORDS END -->\n")
    return "".join(html)

def update_html_file(injected_html):
    if not HTML_PATH.exists():
        print(f"Error: HTML dashboard not found at {HTML_PATH}")
        return False
        
    html_content = HTML_PATH.read_text(encoding="utf-8")
    
    # Check if markers exist
    start_marker = "<!-- EXCEL RECORDS START -->"
    end_marker = "<!-- EXCEL RECORDS END -->"
    
    if start_marker in html_content and end_marker in html_content:
        # Replace existing block
        start_idx = html_content.find(start_marker)
        end_idx = html_content.find(end_marker) + len(end_marker)
        new_content = html_content[:start_idx] + injected_html + html_content[end_idx:]
    else:
        # Suture before footer
        foot_marker = '<div class="foot">'
        if foot_marker in html_content:
            new_content = html_content.replace(foot_marker, injected_html + "\n  " + foot_marker)
        else:
            print("Error: Could not locate insertion point in HTML.")
            return False
            
    HTML_PATH.write_text(new_content, encoding="utf-8")
    print(f"Successfully updated HTML dashboard at {HTML_PATH}")
    return True

def main():
    print("Initializing URC Records Sync to Command Center...")
    if not copy_excel_safely():
        return
        
    wb = load_workbook(TEMP_EXCEL, read_only=True, data_only=True)
    
    # Parse sheets
    apps = parse_app_builds(wb["Apps Builds"])
    accomplishments = parse_accomplishments(wb["Accomplished TW"])
    accounts = parse_new_accounts(wb["New Accounts Added TW"])
    leads = parse_lead_list(wb["Lead List"])
    
    # Parse social posts for Tuesday, Aug 11, 2026 (Day 9, Var 3)
    # Excel date for 2026-08-11 is 46245
    posts = parse_social_posts(wb["Founder_Signal_Challenge_90Day_"], target_date_num=46245)
    
    # Close workbook to release file lock on Windows
    wb.close()
    
    # Build HTML and update file
    injected_html = build_html_sections(apps, accomplishments, accounts, leads, posts)
    update_html_file(injected_html)
    
    # Cleanup
    if TEMP_EXCEL.exists():
        try:
            os.remove(TEMP_EXCEL)
        except Exception as e:
            print(f"Warning: could not remove temp file: {e}")
    print("Sync complete! Your spreadsheet records are now live on your Desktop Command Center.")

if __name__ == "__main__":
    main()
